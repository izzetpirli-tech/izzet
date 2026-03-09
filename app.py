import streamlit as st
import pandas as pd
import json
import os
import io
from datetime import datetime, timedelta
import plotly.express as px
import plotly.graph_objects as go
import psycopg2
import psycopg2.extras
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import hashlib
import requests
import secrets

# ─────────────────────────────────────────────
# SAYFA AYARLARI
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="ELFİGA MANTI - Yönetim Sistemi",
    page_icon="🥟",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─────────────────────────────────────────────
# SABİTLER
# ─────────────────────────────────────────────
KRITIK_SEVIYE = 20.0
SICAKLIK_API_URL = os.environ.get("SICAKLIK_API_URL", "")
TOLERANS = 0.001
VARSAYILAN_RECETE = {
    "Un": 0.6475, "Soğan": 0.100, "Kıyma": 0.046,
    "Yağ": 0.046, "Soya": 0.046, "İrmik": 0.046,
    "Baharat": 0.015, "Tuz": 0.003
}

# ─────────────────────────────────────────────
# CSS TASARIM
# ─────────────────────────────────────────────
st.markdown("""
<style>
    /* Genel arka plan */
    .stApp { background-color: #F5F5F7; }
    
    /* KPI Kartları */
    .kpi-card {
        background: white;
        border-radius: 16px;
        padding: 20px 24px;
        box-shadow: 0 2px 12px rgba(0,0,0,0.07);
        border-left: 5px solid #007AFF;
        margin-bottom: 10px;
    }
    .kpi-card.green { border-left-color: #34C759; }
    .kpi-card.red   { border-left-color: #FF3B30; }
    .kpi-card.orange{ border-left-color: #FF9500; }
    .kpi-label { font-size: 12px; color: #8E8E93; font-weight: 600; margin-bottom: 4px; }
    .kpi-value { font-size: 28px; font-weight: 700; color: #1D1D1F; }
    
    /* Section başlıkları */
    .section-header {
        background: white;
        border-radius: 12px;
        padding: 16px 20px;
        margin-bottom: 16px;
        box-shadow: 0 1px 6px rgba(0,0,0,0.05);
        font-size: 16px;
        font-weight: 700;
        color: #1D1D1F;
    }
    
    /* Butonlar */
    .stButton > button {
        border-radius: 10px !important;
        font-weight: 600 !important;
        border: none !important;
    }
    
    /* Kritik satır */
    .kritik-row { background-color: #FFEEEE !important; color: red !important; }
    
    /* Sidebar */
    .css-1d391kg { background-color: #1D1D1F !important; }
    
    /* Login kutusu */
    .login-box {
        max-width: 400px;
        margin: 60px auto;
        background: white;
        border-radius: 20px;
        padding: 40px;
        box-shadow: 0 8px 32px rgba(0,0,0,0.12);
        text-align: center;
    }
    
    div[data-testid="metric-container"] {
        background: white;
        border-radius: 12px;
        padding: 16px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.06);
    }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# VERİTABANI BAĞLANTISI
# ─────────────────────────────────────────────
@st.cache_resource
def get_db():
    """PostgreSQL bağlantısı - Railway DATABASE_URL kullanır"""
    db_url = os.environ.get("DATABASE_URL")
    if db_url:
        # Railway bazen postgres:// verir, psycopg2 postgresql:// ister
        if db_url.startswith("postgres://"):
            db_url = db_url.replace("postgres://", "postgresql://", 1)
        conn = psycopg2.connect(db_url)
        conn.autocommit = True
        return conn
    else:
        st.error("❌ DATABASE_URL bulunamadı! Railway'de PostgreSQL ekleyin.")
        st.stop()

def init_db():
    """Tabloları oluştur (ilk çalışmada)"""
    conn = get_db()
    cur = conn.cursor()
    
    # Ana veri tablosu - JSON olarak sakla (esneklik için)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS veriler (
            id SERIAL PRIMARY KEY,
            key TEXT UNIQUE NOT NULL,
            value JSONB NOT NULL,
            updated_at TIMESTAMP DEFAULT NOW()
        )
    """)
    
    # Kullanıcılar tablosu
    cur.execute("""
        CREATE TABLE IF NOT EXISTS kullanicilar (
            id SERIAL PRIMARY KEY,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT DEFAULT 'user',
            created_at TIMESTAMP DEFAULT NOW()
        )
    """)
    cur.close()

def db_get(key, default=None):
    """Veritabanından değer oku"""
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT value FROM veriler WHERE key = %s", (key,))
        row = cur.fetchone()
        cur.close()
        return row['value'] if row else default
    except Exception as e:
        st.error(f"DB okuma hatası: {e}")
        return default

def db_set(key, value):
    """Veritabanına değer yaz"""
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO veriler (key, value) VALUES (%s, %s)
            ON CONFLICT (key) DO UPDATE SET value = %s, updated_at = NOW()
        """, (key, json.dumps(value, ensure_ascii=False), json.dumps(value, ensure_ascii=False)))
        cur.close()
        return True
    except Exception as e:
        st.error(f"DB yazma hatası: {e}")
        return False

# ─────────────────────────────────────────────
# VERİ YÖNETİMİ
# ─────────────────────────────────────────────
def veriler_yukle():
    """Tenant verisini yükle — tenant_id varsa tenant tablosundan, yoksa eski tablodan"""
    varsayilan = {
        "stoklar": {k: 0.0 for k in VARSAYILAN_RECETE},
        "detayli_stok": [],
        "hazir_manti_stok": 0.0,
        "hareketler": [],
        "birim_fiyatlar": {k: 0.0 for k in VARSAYILAN_RECETE},
        "coklu_receteler": {
            "Standart Soyalı": {
                "oranlar": VARSAYILAN_RECETE.copy(),
                "etiket_tipi": "Soyalı"
            }
        },
        "aktif_recete_adi": "Standart Soyalı"
    }
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        tenant_id = st.session_state.get("tenant_id")
        if tenant_id:
            cur.execute("SELECT value FROM tenant_veriler WHERE tenant_id = %s AND key = 'ana_veri'", (tenant_id,))
        else:
            cur.execute("SELECT value FROM veriler WHERE key = 'ana_veri'")
        row = cur.fetchone()
        cur.close()
        if row:
            kayitli = row["value"]
            if isinstance(kayitli, str):
                kayitli = json.loads(kayitli)
            for k, v in varsayilan.items():
                if k not in kayitli:
                    kayitli[k] = v
            return kayitli
    except Exception as e:
        st.error(f"Veri yükleme hatası: {e}")
    return varsayilan

def veriler_kaydet(veriler):
    """Tenant verisini kaydet"""
    try:
        conn = get_db()
        cur = conn.cursor()
        veri_json = json.dumps(veriler, ensure_ascii=False)
        tenant_id = st.session_state.get("tenant_id")
        if tenant_id:
            cur.execute("""
                INSERT INTO tenant_veriler (tenant_id, key, value)
                VALUES (%s, 'ana_veri', %s::jsonb)
                ON CONFLICT (tenant_id, key) DO UPDATE SET value = %s::jsonb, updated_at = NOW()
            """, (tenant_id, veri_json, veri_json))
        else:
            cur.execute("""
                INSERT INTO veriler (key, value) VALUES ('ana_veri', %s::jsonb)
                ON CONFLICT (key) DO UPDATE SET value = %s::jsonb, updated_at = NOW()
            """, (veri_json, veri_json))
        cur.close()
        st.session_state.veriler = veriler
        return True
    except Exception as e:
        st.error(f"Kaydetme hatası: {e}")
        return False

def sayiya_cevir(s):
    try:
        return float(str(s).replace(",", ".")) if s else None
    except:
        return None

# ─────────────────────────────────────────────
# KİMLİK DOĞRULAMA
# ─────────────────────────────────────────────
def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def kullanici_dogrula(username, password):
    """Önce tenant tablosunda ara, yoksa eski tabloda"""
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        
        # Tenant sistemi: tüm tenantlarda bu kullanıcıyı ara
        cur.execute("""
            SELECT tk.*, t.id as tenant_id, t.firma_adi, t.slug, t.durum, t.bitis_tarihi
            FROM tenant_kullanicilar tk
            JOIN tenants t ON tk.tenant_id = t.id
            WHERE tk.username = %s AND tk.password_hash = %s
        """, (username, hash_password(password)))
        user = cur.fetchone()
        
        if user:
            # Abonelik kontrolü
            bugun = datetime.now().date()
            if user["durum"] == "pasif":
                cur.close()
                return None, "Hesabınız pasif edilmiştir. Lütfen yöneticinize başvurun."
            if user["bitis_tarihi"] and user["bitis_tarihi"] < bugun:
                cur.close()
                return None, "Abonelik süreniz dolmuştur. Lütfen yenileyin."
            cur.close()
            return dict(user), None
        
        # Eski sistem (geriye dönük uyumluluk)
        cur.execute("SELECT * FROM kullanicilar WHERE username = %s", (username,))
        old_user = cur.fetchone()
        cur.close()
        if old_user and old_user["password_hash"] == hash_password(password):
            return dict(old_user), None
        
        return None, "Kullanıcı adı veya parola hatalı."
    except Exception as e:
        return None, f"Bağlantı hatası: {e}"

def admin_kullanici_olustur():
    """İlk çalıştırmada eski sistem için admin oluştur (geriye dönük)"""
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS kullanicilar (
                id SERIAL PRIMARY KEY,
                username TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                role TEXT DEFAULT 'user',
                created_at TIMESTAMP DEFAULT NOW()
            )
        """)
        cur.execute("SELECT COUNT(*) FROM kullanicilar WHERE username = 'admin'")
        count = cur.fetchone()[0]
        if count == 0:
            admin_pass = os.environ.get("ADMIN_PASSWORD", "elfiga2024")
            cur.execute(
                "INSERT INTO kullanicilar (username, password_hash, role) VALUES (%s, %s, %s)",
                ("admin", hash_password(admin_pass), "admin")
            )
        cur.close()
    except Exception as e:
        pass

def login_sayfasi():
    """Login ekranı"""
    col1, col2, col3 = st.columns([1, 1.5, 1])
    with col2:
        st.markdown("""
        <div style='text-align:center; padding: 40px 0 20px 0;'>
            <div style='font-size: 64px;'>🥟</div>
            <h1 style='color: #1D1D1F; font-size: 28px; margin: 10px 0 4px 0;'>ELFİGA MANTI</h1>
            <p style='color: #8E8E93; font-size: 14px;'>Yönetim Sistemi</p>
        </div>
        """, unsafe_allow_html=True)
        
        with st.form("login_form"):
            username = st.text_input("👤 Kullanıcı Adı", placeholder="kullanıcı adınız")
            password = st.text_input("🔒 Parola", type="password", placeholder="••••••••")
            submit = st.form_submit_button("Giriş Yap", use_container_width=True, type="primary")
            
            if submit:
                user, hata = kullanici_dogrula(username, password)
                if user:
                    st.session_state.logged_in = True
                    st.session_state.username = username
                    st.session_state.role = user.get('rol', user.get('role', 'user'))
                    if 'tenant_id' in user:
                        st.session_state.tenant_id = user['tenant_id']
                        st.session_state.firma_adi = user.get('firma_adi', 'ELFİGA MANTI')
                    st.rerun()
                else:
                    st.error(f"❌ {hata}")

# ─────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────
def sidebar_menu():
    with st.sidebar:
        st.markdown("""
        <div style='text-align:center; padding: 20px 0;'>
            <div style='font-size: 40px;'>🥟</div>
            <h2 style='color: white; margin: 8px 0 2px 0; font-size: 16px;'>ELFİGA MANTI</h2>
            <p style='color: #8E8E93; font-size: 11px; margin: 0;'>Yönetim Sistemi</p>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("---")
        
        menu = st.radio(
            "📍 Menü",
            ["📊 Patron Ekranı", "🏭 Üretim", "📍 Depom", "📦 Depo & Giriş", "🌡️ Sıcaklık", "⚙️ Ayarlar"],
            label_visibility="collapsed"
        )
        
        st.markdown("---")
        st.markdown(f"👤 **{st.session_state.get('username', '')}**")
        firma = st.session_state.get('firma_adi', 'ELFİGA MANTI')
        st.markdown(f"🏢 *{firma}*")
        if st.button("🚪 Çıkış", use_container_width=True):
            st.session_state.logged_in = False
            st.session_state.username = ""
            st.rerun()
    
    return menu

# ─────────────────────────────────────────────
# SAYFA: PATRON EKRANI (DASHBOARD)
# ─────────────────────────────────────────────
def patron_ekrani(v):
    st.markdown("## 📊 Patron Ekranı")
    
    st.markdown("---")
    
    # KPI Kartları
    st.markdown("### 📈 Genel Durum")
    k1, k2, k3, k4 = st.columns(4)
    
    stoklar = v["stoklar"]
    fiyatlar = v["birim_fiyatlar"]
    toplam_deger = sum(stoklar.get(m, 0) * fiyatlar.get(m, 0) for m in stoklar)
    kritik_sayisi = sum(1 for k in stoklar.values() if k < KRITIK_SEVIYE)
    
    current_month = datetime.now().strftime("%m.%Y")
    aylik_uretim = sum(
        h.get("miktar", 0) for h in v["hareketler"]
        if h.get("islem") == "Üretim" and h.get("tarih", "").endswith(current_month)
    )
    
    with k1:
        st.metric("💰 Toplam Stok Değeri", f"{toplam_deger:,.2f} TL")
    with k2:
        st.metric("🥟 Hazır Mantı", f"{v['hazir_manti_stok']:.2f} KG")
    with k3:
        st.metric("🏭 Bu Ay Üretim", f"{aylik_uretim:.2f} KG")
    with k4:
        st.metric("⚠️ Kritik Stok", f"{kritik_sayisi} Kalem", delta_color="inverse")
    
    st.markdown("---")
    
    # Grafikler
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("### 🥧 Stok Değeri Dağılımı")
        degerler = [(m, stoklar.get(m, 0) * fiyatlar.get(m, 0)) for m in stoklar if stoklar.get(m, 0) * fiyatlar.get(m, 0) > 0]
        if degerler:
            df_pie = pd.DataFrame(degerler, columns=["Malzeme", "Değer (TL)"])
            fig = px.pie(df_pie, values="Değer (TL)", names="Malzeme",
                        color_discrete_sequence=px.colors.qualitative.Set3)
            fig.update_layout(margin=dict(t=20, b=20, l=20, r=20), height=300)
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("Stok verisi yok.")
    
    with col2:
        st.markdown("### 📦 En Çok Stok (KG)")
        stok_sirali = sorted([(m, k) for m, k in stoklar.items() if k > 0], key=lambda x: x[1], reverse=True)[:8]
        if stok_sirali:
            df_bar = pd.DataFrame(stok_sirali, columns=["Malzeme", "Miktar (KG)"])
            fig2 = px.bar(df_bar, x="Malzeme", y="Miktar (KG)",
                         color="Miktar (KG)", color_continuous_scale="Blues",
                         text_auto=".1f")
            fig2.update_layout(margin=dict(t=20, b=20, l=20, r=20), height=300, showlegend=False)
            st.plotly_chart(fig2, use_container_width=True)
        else:
            st.info("Stok verisi yok.")
    
    # Son hareketler
    st.markdown("### 🕐 Son 10 Hareket")
    if v["hareketler"]:
        df_son = pd.DataFrame(v["hareketler"][:10])
        df_son["miktar"] = df_son["miktar"].apply(lambda x: abs(float(x)))
        st.dataframe(
            df_son[["tarih", "malzeme", "miktar", "islem", "parti"]].rename(
                columns={"tarih": "Tarih", "malzeme": "Malzeme/Ürün",
                         "miktar": "Miktar (KG)", "islem": "İşlem", "parti": "Parti No"}
            ),
            use_container_width=True, hide_index=True
        )

# ─────────────────────────────────────────────
# SAYFA: ÜRETİM
# ─────────────────────────────────────────────
def uretim_sayfasi(v):
    st.markdown("## 🏭 Üretim Yönetimi")
    
    col_info, col_satis = st.columns([2, 1])
    with col_info:
        stok_rengi = "green" if v["hazir_manti_stok"] > 50 else "orange" if v["hazir_manti_stok"] > 10 else "red"
        st.markdown(f"""
        <div class='kpi-card {"green" if v["hazir_manti_stok"] > 50 else "orange"}'>
            <div class='kpi-label'>HAZIR MANTI STOĞU</div>
            <div class='kpi-value'>{v["hazir_manti_stok"]:.2f} KG</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col_satis:
        with st.form("satis_form"):
            st.markdown("**🛒 Satış Düş**")
            satis_kg = st.number_input("Satış Miktarı (KG)", min_value=0.0, step=0.5)
            if st.form_submit_button("SATIŞ KAYDET", type="primary"):
                if satis_kg > 0 and satis_kg <= v["hazir_manti_stok"]:
                    v["hazir_manti_stok"] -= satis_kg
                    v["hareketler"].insert(0, {
                        "tarih": datetime.now().strftime("%d.%m.%Y"),
                        "malzeme": "SATIŞ", "miktar": -satis_kg,
                        "fiyat": 0, "parti": "-", "fatura": "-", "islem": "Satış"
                    })
                    veriler_kaydet(v)
                    st.success(f"✅ {satis_kg:.2f} KG satış kaydedildi!")
                    st.rerun()
                else:
                    st.error("Geçersiz miktar veya yetersiz stok!")
    
    st.markdown("---")
    
    tab1, tab2 = st.tabs(["✅ Üretim Başlat", "📋 Sipariş Planla"])
    
    with tab1:
        col_form, col_stok = st.columns([1, 2])
        
        with col_form:
            st.markdown("#### İmalat Kontrol")
            recete_keys = list(v["coklu_receteler"].keys())
            secilen_recete = st.selectbox("📋 Reçete Seç", recete_keys,
                                          index=recete_keys.index(v["aktif_recete_adi"]) if v["aktif_recete_adi"] in recete_keys else 0)
            
            if secilen_recete != v["aktif_recete_adi"]:
                v["aktif_recete_adi"] = secilen_recete
                veriler_kaydet(v)
            
            aktif_recete = v["coklu_receteler"][secilen_recete]["oranlar"]
            
            # Kapasite analizi
            max_uretim = 99999
            kisitlayan = ""
            for m, o in aktif_recete.items():
                if o > 0:
                    stok = max(0, v["stoklar"].get(m, 0))
                    kapasite = stok / o
                    if kapasite < max_uretim:
                        max_uretim = kapasite
                        kisitlayan = m
            
            if max_uretim < 99999:
                st.info(f"📊 Max Üretilebilir: **{max_uretim:.2f} KG** (kısıt: {kisitlayan})")
            
            with st.form("uretim_form"):
                uretim_kg = st.number_input("Üretim Miktarı (KG)", min_value=0.1, step=1.0, value=10.0)
                uret_btn = st.form_submit_button("✅ ÜRET", type="primary", use_container_width=True)
                
                if uret_btn:
                    # Stok kontrol
                    yetersiz = []
                    for m, o in aktif_recete.items():
                        if o > 0 and v["stoklar"].get(m, 0) + TOLERANS < uretim_kg * o:
                            yetersiz.append(f"{m} (eksik: {uretim_kg*o - v['stoklar'].get(m,0):.2f} KG)")
                    
                    if yetersiz:
                        st.error("❌ Yetersiz stok:\n" + "\n".join(yetersiz))
                    else:
                        kullanilan = []
                        uretim_parti = "URT-" + datetime.now().strftime("%y%m%d-%H%M")
                        
                        for m, oran in aktif_recete.items():
                            if oran == 0:
                                continue
                            dusulecek = uretim_kg * oran
                            temp = 0
                            mevcut = sorted(
                                [p for p in v["detayli_stok"] if p["malzeme"] == m and p["kalan"] > TOLERANS],
                                key=lambda x: datetime.strptime(x["tarih"], "%d.%m.%Y")
                            )
                            for parti in mevcut:
                                if temp >= dusulecek:
                                    break
                                lazim = dusulecek - temp
                                take = min(lazim, parti["kalan"])
                                for d in v["detayli_stok"]:
                                    if d["parti"] == parti["parti"] and d["malzeme"] == m:
                                        d["kalan"] -= take
                                        if d["kalan"] < TOLERANS:
                                            d["kalan"] = 0.0
                                        break
                                temp += take
                                kullanilan.append({"malzeme": m, "miktar": take, "parti": parti["parti"], "fatura": parti["fatura"]})
                            
                            v["stoklar"][m] -= dusulecek
                            if v["stoklar"][m] < TOLERANS:
                                v["stoklar"][m] = 0.0
                        
                        v["hazir_manti_stok"] += uretim_kg
                        v["hareketler"].insert(0, {
                            "tarih": datetime.now().strftime("%d.%m.%Y"),
                            "malzeme": f"ÜRETİM ({secilen_recete})",
                            "miktar": uretim_kg, "fiyat": 0,
                            "fatura": "-", "parti": uretim_parti,
                            "islem": "Üretim",
                            "kullanilan_detay": kullanilan
                        })
                        veriler_kaydet(v)
                        st.success(f"✅ {uretim_kg:.2f} KG üretim tamamlandı! Parti: {uretim_parti}")
                        
                        # Etiket bilgileri
                        etiket_tipi = v["coklu_receteler"][secilen_recete].get("etiket_tipi", "Soyalı")
                        skt = (datetime.now() + timedelta(days=180)).strftime("%d.%m.%Y")
                        st.info(f"""
                        **📋 ETİKET BİLGİLERİ**
                        - Ürün: MANTI ({etiket_tipi})
                        - Parti No: {uretim_parti}
                        - Üretim Tarihi: {datetime.now().strftime("%d.%m.%Y")}
                        - SKT: {skt}
                        """)
                        st.rerun()
        
        with col_stok:
            st.markdown("#### 📦 Hammadde Stoku")
            stok_data = []
            for m, k in v["stoklar"].items():
                fiyat = v["birim_fiyatlar"].get(m, 0)
                stok_data.append({
                    "Malzeme": m,
                    "Stok (KG)": round(k, 2),
                    "Birim Fiyat (TL)": round(fiyat, 2),
                    "Durum": "⚠️ KRİTİK" if k < KRITIK_SEVIYE else "✅ Normal"
                })
            df_stok = pd.DataFrame(stok_data)
            st.dataframe(df_stok, use_container_width=True, hide_index=True,
                        column_config={"Stok (KG)": st.column_config.NumberColumn(format="%.2f")})
    
    with tab2:
        st.markdown("#### 📋 Sipariş / Planlama Sihirbazı")
        hedef = st.number_input("Hedef Üretim Miktarı (KG)", min_value=1.0, step=10.0, value=100.0)
        
        aktif_r = v["coklu_receteler"][v["aktif_recete_adi"]]["oranlar"]
        plan_data = []
        for m, o in aktif_r.items():
            if o > 0:
                gereken = hedef * o
                mevcut = v["stoklar"].get(m, 0)
                fark = mevcut - gereken
                plan_data.append({
                    "Malzeme": m,
                    "Gereken (KG)": round(gereken, 2),
                    "Mevcut Stok (KG)": round(mevcut, 2),
                    "Fark (KG)": round(fark, 2),
                    "Durum": "✅ Yeterli" if fark >= -TOLERANS else f"❌ Eksik ({abs(fark):.2f} KG)"
                })
        
        df_plan = pd.DataFrame(plan_data)
        st.dataframe(df_plan, use_container_width=True, hide_index=True)

# ─────────────────────────────────────────────
# SAYFA: DEPOM (ANLIK)
# ─────────────────────────────────────────────
def depom_sayfasi(v):
    st.markdown("## 📍 Depom (Anlık Durum)")

    st.markdown(f"""
    <div class='kpi-card green'>
        <div class='kpi-label'>HAZIR MANTI STOĞU (TOPLAM)</div>
        <div class='kpi-value'>{v["hazir_manti_stok"]:.2f} KG</div>
    </div>
    """, unsafe_allow_html=True)

    # ── ÜRETİM PARTİLERİ + İZLENEBİLİRLİK ──────────────────────────────
    st.markdown("#### 🥟 Üretim Partileri — Birini seçerek detay & izlenebilirlik formunu aç")
    manti_h = [h for h in v["hareketler"] if h.get("islem") == "Üretim"]

    if not manti_h:
        st.info("Üretim kaydı bulunmuyor.")
    else:
        parti_secenekler = [
            f"{h.get('parti')} | {h.get('miktar')} KG | {h.get('tarih')} | {'✅ Sevk' if h.get('sevkiyat_detay') else '⏳ Bekliyor'}"
            for h in manti_h
        ]
        secim = st.selectbox("📦 Parti Seç", ["— Seçiniz —"] + parti_secenekler)

        if secim != "— Seçiniz —":
            idx = parti_secenekler.index(secim)
            hedef = manti_h[idx]
            hedef_global_idx = next(
                (i for i, h in enumerate(v["hareketler"])
                 if h.get("parti") == hedef.get("parti") and h.get("islem") == "Üretim"), None
            )

            st.markdown("---")
            st.markdown(f"### 🔍 {hedef['parti']} — İzlenebilirlik Formu")

            col_sol, col_sag = st.columns(2)

            with col_sol:
                st.markdown("**🧪 Kullanılan Malzemeler**")
                kullanilan = hedef.get("kullanilan_detay", [])
                if not kullanilan:
                    recete_adi = hedef.get("malzeme", "").replace("ÜRETİM (", "").replace(")", "")
                    recete_data = v["coklu_receteler"].get(recete_adi) or                                   v["coklu_receteler"].get(v["aktif_recete_adi"], {"oranlar": {}})
                    miktar = hedef.get("miktar", 0)
                    kullanilan = [
                        {"malzeme": m, "miktar": round(miktar * o, 3), "parti": "GEÇMİŞ", "fatura": "-"}
                        for m, o in recete_data["oranlar"].items() if o > 0
                    ]
                    st.caption("⚠️ Geçmiş kayıt — teorik hesaplama gösteriliyor")

                df_kul = pd.DataFrame([{
                    "Malzeme": k["malzeme"],
                    "Miktar (KG)": round(k["miktar"], 3),
                    "Parti": k.get("parti", "-"),
                    "Fatura": k.get("fatura", "-")
                } for k in kullanilan])
                st.dataframe(df_kul, use_container_width=True, hide_index=True)

                st.markdown("**🏷️ Etiket Bilgileri**")
                recete_adi2 = hedef.get("malzeme", "").replace("ÜRETİM (", "").replace(")", "")
                recete_data2 = v["coklu_receteler"].get(recete_adi2) or                                v["coklu_receteler"].get(v["aktif_recete_adi"], {"etiket_tipi": "Soyalı"})
                etiket_tipi = recete_data2.get("etiket_tipi", "Soyalı")
                try:
                    uretim_dt = datetime.strptime(hedef["tarih"], "%d.%m.%Y")
                except:
                    uretim_dt = datetime.now()
                skt = (uretim_dt + timedelta(days=180)).strftime("%d.%m.%Y")
                paket_secenekler = [0.25, 0.5, 1.0, 2.0, 5.0, 10.0]
                paket_kg = st.selectbox("Paket Ağırlığı (KG)", paket_secenekler, index=4, key=f"paket_{hedef['parti']}")
                kopya = int(hedef.get("miktar", 0) / paket_kg) if paket_kg > 0 else 0
                agirlik_str = f"{int(paket_kg*1000)} GR" if paket_kg < 1 else f"{int(paket_kg)} KG"
                st.markdown(f"""
<div style="background:#f0f7ff;border:1px solid #007AFF;border-radius:10px;padding:14px;margin-top:8px;line-height:2.2">
<div style="display:flex;justify-content:space-between;border-bottom:1px solid #d0e4ff;padding-bottom:6px;margin-bottom:6px">
  <span style="color:#555;font-size:13px">ÜRÜN</span>
  <span style="font-weight:700;color:#1D1D1F">MANTI ({etiket_tipi})</span>
</div>
<div style="display:flex;justify-content:space-between;border-bottom:1px solid #d0e4ff;padding-bottom:6px;margin-bottom:6px">
  <span style="color:#555;font-size:13px">AĞIRLIK</span>
  <span style="font-weight:700;color:#1D1D1F">{agirlik_str}</span>
</div>
<div style="display:flex;justify-content:space-between;border-bottom:1px solid #d0e4ff;padding-bottom:6px;margin-bottom:6px">
  <span style="color:#555;font-size:13px">PARTİ NO</span>
  <span style="font-weight:700;color:#007AFF">{hedef["parti"]}</span>
</div>
<div style="display:flex;justify-content:space-between;border-bottom:1px solid #d0e4ff;padding-bottom:6px;margin-bottom:6px">
  <span style="color:#555;font-size:13px">ÜRETİM TARİHİ</span>
  <span style="font-weight:700;color:#1D1D1F">{hedef["tarih"]}</span>
</div>
<div style="display:flex;justify-content:space-between;border-bottom:1px solid #d0e4ff;padding-bottom:6px;margin-bottom:6px">
  <span style="color:#555;font-size:13px">SKT</span>
  <span style="font-weight:700;color:#FF3B30">{skt}</span>
</div>
<div style="display:flex;justify-content:space-between">
  <span style="color:#555;font-size:13px">KOPYA ADEDİ</span>
  <span style="font-weight:700;color:#34C759">{kopya} adet</span>
</div>
</div>
""", unsafe_allow_html=True)

            with col_sag:
                st.markdown("**🚚 Sevkiyat Takibi**")
                fire = st.slider("Fire Oranı (%)", 0, 20, int(float(hedef.get("fire_orani", 4))), key=f"fire_{hedef['parti']}")
                net = hedef["miktar"] * (1 - fire / 100)
                st.metric("Net Dağıtılabilir", f"{net:.2f} KG")
                notlar = st.text_input("Notlar", value=hedef.get("uretim_notu", ""), key=f"not_{hedef['parti']}")

                # Excel kopyala-yapıştır alanı
                st.markdown("**📋 Excel'den Yapıştır** *(Tarih → Firma → KG → Fatura sırası)*")
                paste_text = st.text_area(
                    "Excel'den kopyaladığın satırları buraya yapıştır",
                    height=100,
                    placeholder="01.03.2026\tFirma A\t50\tFAT-001",
                    key=f"paste_{hedef['parti']}"
                )

                with st.form(f"sevk_form_{hedef['parti']}"):
                    st.markdown("**Veya tek satır ekle:**")
                    c1, c2 = st.columns(2)
                    with c1:
                        s_tarih = st.date_input("Tarih")
                        s_firma = st.text_input("Firma Adı")
                    with c2:
                        s_miktar = st.number_input("Miktar (KG)", min_value=0.0, step=0.5)
                        s_fatura = st.text_input("Fatura No")
                    col_b1, col_b2, col_b3 = st.columns(3)
                    with col_b1:
                        excel_aktar = st.form_submit_button("📋 Excel'i Aktar", type="primary")
                    with col_b2:
                        ekle = st.form_submit_button("➕ Tek Ekle")
                    with col_b3:
                        sadece_kaydet = st.form_submit_button("💾 Kaydet")

                    mevcut_sevk = list(hedef.get("sevkiyat_detay", []))

                    if excel_aktar and paste_text.strip():
                        satirlar = paste_text.strip().split("\n")
                        eklenen = 0
                        hatalar = []
                        for satir in satirlar:
                            satir = satir.strip()
                            if not satir:
                                continue
                            # Tab veya birden fazla boşluk ile ayır
                            kolonlar = [k.strip() for k in satir.replace("\t", "\t").split("\t")]
                            if len(kolonlar) < 3:
                                kolonlar = [k.strip() for k in satir.split() if k.strip()]
                            if len(kolonlar) >= 3:
                                try:
                                    tarih_col = kolonlar[0]
                                    firma_col = kolonlar[1]
                                    kg_col = float(str(kolonlar[2]).replace(",", "."))
                                    fat_col = kolonlar[3] if len(kolonlar) > 3 else "-"
                                    mevcut_sevk.append({"tarih": tarih_col, "firma": firma_col, "miktar": kg_col, "fatura": fat_col})
                                    eklenen += 1
                                except Exception as ex:
                                    hatalar.append(satir)
                        v["hareketler"][hedef_global_idx]["sevkiyat_detay"] = mevcut_sevk
                        v["hareketler"][hedef_global_idx]["fire_orani"] = str(fire)
                        v["hareketler"][hedef_global_idx]["uretim_notu"] = notlar
                        veriler_kaydet(v)
                        if eklenen > 0:
                            st.success(f"✅ {eklenen} satır aktarıldı!")
                        if hatalar:
                            st.warning(f"⚠️ Şu satırlar okunamadı: {hatalar}")
                        st.rerun()

                    if ekle and s_firma and s_miktar > 0:
                        mevcut_sevk.append({"tarih": s_tarih.strftime("%d.%m.%Y"), "firma": s_firma, "miktar": s_miktar, "fatura": s_fatura})
                        v["hareketler"][hedef_global_idx]["sevkiyat_detay"] = mevcut_sevk
                        v["hareketler"][hedef_global_idx]["fire_orani"] = str(fire)
                        v["hareketler"][hedef_global_idx]["uretim_notu"] = notlar
                        veriler_kaydet(v)
                        st.success("✅ Eklendi!")
                        st.rerun()
                    if sadece_kaydet:
                        v["hareketler"][hedef_global_idx]["fire_orani"] = str(fire)
                        v["hareketler"][hedef_global_idx]["uretim_notu"] = notlar
                        v["hareketler"][hedef_global_idx]["sevkiyat_detay"] = mevcut_sevk
                        veriler_kaydet(v)
                        st.success("✅ Kaydedildi!")
                        st.rerun()

                mevcut_sevk = hedef.get("sevkiyat_detay", [])
                if mevcut_sevk:
                    st.markdown("**Mevcut Sevkiyatlar:**")
                    df_sevk = pd.DataFrame(mevcut_sevk)
                    st.dataframe(df_sevk, use_container_width=True, hide_index=True)
                    toplam_sevk = sum(float(s["miktar"]) for s in mevcut_sevk)
                    st.metric("Toplam Sevkiyat", f"{toplam_sevk:.2f} KG")
                    if st.button("🖨️ Sevkiyat Raporu İndir", key=f"rapor_{hedef['parti']}"):
                        html = f"""<html><head><meta charset='utf-8'><style>
                        body{{font-family:Arial,sans-serif;font-size:13px;padding:30px}}
                        h2{{color:#007AFF;border-bottom:2px solid #007AFF;padding-bottom:8px}}
                        table{{width:100%;border-collapse:collapse;margin:15px 0}}
                        th,td{{border:1px solid #ddd;padding:8px}}th{{background:#f5f5f7;font-weight:bold}}
                        .box{{background:#f9f9f9;border:1px solid #ddd;padding:12px;border-radius:8px;margin:10px 0}}
                        </style></head><body>
                        <h2>ELFİGA MANTI — Sevkiyat Raporu</h2>
                        <div class='box'><b>Ürün:</b> {hedef["malzeme"]}<br><b>Parti No:</b> {hedef["parti"]}<br>
                        <b>Üretim Miktarı:</b> {hedef["miktar"]} KG<br><b>Fire Oranı:</b> %{fire}<br>
                        <b>Net Dağıtılabilir:</b> {net:.2f} KG<br><b>Not:</b> {notlar}</div>
                        <h3>Kullanılan Malzemeler</h3>
                        <table><tr><th>Malzeme</th><th>KG</th><th>Parti</th><th>Fatura</th></tr>
                        {"".join(f"<tr><td>{k['malzeme']}</td><td>{k['miktar']:.2f}</td><td>{k.get('parti','-')}</td><td>{k.get('fatura','-')}</td></tr>" for k in kullanilan)}
                        </table><h3>Sevkiyat Detayları</h3>
                        <table><tr><th>Tarih</th><th>Firma</th><th>Miktar (KG)</th><th>Fatura</th></tr>
                        {"".join(f"<tr><td>{s['tarih']}</td><td>{s['firma']}</td><td>{s['miktar']}</td><td>{s['fatura']}</td></tr>" for s in mevcut_sevk)}
                        <tr><td colspan='2'><b>TOPLAM</b></td><td><b>{toplam_sevk:.2f}</b></td><td></td></tr></table>
                        <div style='display:flex;justify-content:space-between;margin-top:60px'>
                        <div style='border-top:1px solid #000;padding-top:8px;width:200px;text-align:center'>Üretim Sorumlusu</div>
                        <div style='border-top:1px solid #000;padding-top:8px;width:200px;text-align:center'>Onay</div>
                        </div></body></html>"""
                        st.download_button("⬇️ Raporu İndir (HTML)", data=html.encode("utf-8"),
                                           file_name=f"sevkiyat_{hedef['parti']}.html",
                                           mime="text/html", key=f"dl_{hedef['parti']}")

    st.markdown("---")
    st.markdown("#### 📦 Kalan Hammadde Stoku (Parti Bazlı)")
    kalan = [d for d in v["detayli_stok"] if d.get("kalan", 0) > TOLERANS]
    if kalan:
        df_k = pd.DataFrame([{
            "Malzeme": d["malzeme"], "Parti No": d["parti"],
            "Giriş (KG)": round(d.get("miktar", 0), 2),
            "Kalan (KG)": round(d.get("kalan", 0), 2), "Tarih": d["tarih"]
        } for d in sorted(kalan, key=lambda x: x.get("kalan", 0), reverse=True)])
        st.dataframe(df_k, use_container_width=True, hide_index=True)
        st.markdown("**🗑️ Fire / Zayi Bildir**")
        parti_listesi = [f"{d['malzeme']} - {d['parti']}" for d in kalan]
        secili_parti = st.selectbox("Parti Seç", parti_listesi, key="zayi_secim")
        if st.button("🗑️ Seçili Partiyi Sıfırla", type="secondary"):
            for d in v["detayli_stok"]:
                if f"{d['malzeme']} - {d['parti']}" == secili_parti:
                    kalan_miktar = d["kalan"]
                    v["stoklar"][d["malzeme"]] = max(0, v["stoklar"].get(d["malzeme"], 0) - kalan_miktar)
                    v["hareketler"].insert(0, {"tarih": datetime.now().strftime("%d.%m.%Y"),
                        "malzeme": f"{d['malzeme']} (FİRE/ZAYİ)", "miktar": -kalan_miktar,
                        "fiyat": 0, "parti": d["parti"], "fatura": "-", "islem": "Zayi"})
                    d["kalan"] = 0.0
                    break
            veriler_kaydet(v)
            st.success("✅ Parti sıfırlandı!")
            st.rerun()
    else:
        st.info("Aktif parti bulunamadı.")

# ─────────────────────────────────────────────
# SAYFA: DEPO & GİRİŞ
# ─────────────────────────────────────────────
def depo_giris_sayfasi(v):
    st.markdown("## 📦 Depo & Malzeme Girişi")
    
    tab1, tab2, tab3 = st.tabs(["➕ Malzeme Girişi", "📋 Hareket Kayıtları", "🚚 Sevkiyat"])
    
    with tab1:
        col_form, col_giris = st.columns([1, 1])
        
        with col_form:
            st.markdown("#### Malzeme Girişi")
            with st.form("stok_giris_form", clear_on_submit=True):
                malzeme_listesi = sorted(list(v["stoklar"].keys()))
                yeni_mlz_sec = st.selectbox("Malzeme", malzeme_listesi + ["+ Yeni Malzeme Ekle"])
                
                yeni_mlz_adi = ""
                if yeni_mlz_sec == "+ Yeni Malzeme Ekle":
                    yeni_mlz_adi = st.text_input("Yeni Malzeme Adı")
                
                col_a, col_b = st.columns(2)
                with col_a:
                    miktar = st.number_input("Miktar (KG)", min_value=0.001, step=0.1)
                    tarih = st.date_input("Tarih", value=datetime.now())
                with col_b:
                    fiyat = st.number_input("Fiyat (TL/KG)", min_value=0.0, step=0.1)
                    parti = st.text_input("Parti No", placeholder="Boş bırakırsanız otomatik")
                
                fatura = st.text_input("Fatura No")
                
                submit = st.form_submit_button("💾 KAYDET", type="primary", use_container_width=True)
                
                if submit:
                    malzeme = yeni_mlz_adi if yeni_mlz_sec == "+ Yeni Malzeme Ekle" else yeni_mlz_sec
                    
                    if not malzeme:
                        st.error("Malzeme adı boş olamaz!")
                    else:
                        if malzeme not in v["stoklar"]:
                            v["stoklar"][malzeme] = 0.0
                            v["birim_fiyatlar"][malzeme] = 0.0
                            for r in v["coklu_receteler"].values():
                                r["oranlar"][malzeme] = 0.0
                        
                        parti_no = parti if parti else "PARTI-" + datetime.now().strftime("%y%m%d%H%M")
                        tarih_str = tarih.strftime("%d.%m.%Y")
                        
                        v["stoklar"][malzeme] += miktar
                        v["birim_fiyatlar"][malzeme] = fiyat
                        v["detayli_stok"].insert(0, {
                            "malzeme": malzeme, "miktar": miktar, "kalan": miktar,
                            "fatura": fatura, "parti": parti_no, "tarih": tarih_str
                        })
                        v["hareketler"].insert(0, {
                            "tarih": tarih_str, "malzeme": malzeme, "miktar": miktar,
                            "fiyat": fiyat, "parti": parti_no, "fatura": fatura, "islem": "Giriş"
                        })
                        veriler_kaydet(v)
                        st.success(f"✅ {malzeme} - {miktar:.2f} KG kaydedildi!")
                        st.rerun()
        
        with col_giris:
            st.markdown("#### 🌶️ Baharat Karışımı Oluştur")
            with st.expander("Karışım Oluştur"):
                cikis_mlz = st.selectbox("Çıkış Malzemesi", list(v["stoklar"].keys()), key="karisim_cikis")
                
                karisim_data = []
                st.markdown("Karışıma eklenecek malzemeler:")
                for i in range(5):
                    col_m, col_kg = st.columns(2)
                    with col_m:
                        m = st.selectbox(f"Malzeme {i+1}", ["Seçiniz"] + list(v["stoklar"].keys()), key=f"kar_m_{i}")
                    with col_kg:
                        kg = st.number_input(f"KG {i+1}", min_value=0.0, key=f"kar_kg_{i}")
                    if m != "Seçiniz" and kg > 0:
                        karisim_data.append((m, kg))
                
                if st.button("✅ Karışımı Kaydet") and karisim_data:
                    toplam = sum(kg for _, kg in karisim_data)
                    for m, kg in karisim_data:
                        v["stoklar"][m] -= kg
                    v["stoklar"][cikis_mlz] += toplam
                    kp = "KRS-" + datetime.now().strftime("%y%m%d")
                    v["hareketler"].insert(0, {
                        "tarih": datetime.now().strftime("%d.%m.%Y"),
                        "malzeme": cikis_mlz, "miktar": toplam,
                        "fiyat": 0, "parti": kp, "fatura": "-", "islem": "Karışım"
                    })
                    veriler_kaydet(v)
                    st.success(f"✅ {toplam:.2f} KG karışım oluşturuldu!")
                    st.rerun()
    
    with tab2:
        st.markdown("#### 📋 Hareket Kayıtları")
        
        # Filtreler
        col_f1, col_f2, col_f3 = st.columns(3)
        with col_f1:
            filtre_islem = st.multiselect("İşlem Tipi", ["Giriş", "Üretim", "Satış", "Zayi", "Karışım"], default=["Giriş", "Üretim", "Satış"])
        with col_f2:
            filtre_mlz = st.text_input("Malzeme/Ürün Ara", "")
        with col_f3:
            filtre_parti = st.text_input("Parti No Ara", "")
        
        hareketler = v["hareketler"]
        if filtre_islem:
            hareketler = [h for h in hareketler if h.get("islem") in filtre_islem]
        if filtre_mlz:
            hareketler = [h for h in hareketler if filtre_mlz.lower() in h.get("malzeme", "").lower()]
        if filtre_parti:
            hareketler = [h for h in hareketler if filtre_parti.lower() in h.get("parti", "").lower()]
        
        if hareketler:
            df_h = pd.DataFrame([{
                "Tarih": h.get("tarih"),
                "Malzeme/Ürün": h.get("malzeme"),
                "Miktar (KG)": abs(h.get("miktar", 0)),
                "Fiyat (TL)": h.get("fiyat", 0),
                "Parti No": h.get("parti"),
                "Fatura No": h.get("fatura"),
                "İşlem": h.get("islem"),
                "Sevkiyat": "✅" if h.get("sevkiyat_detay") else ""
            } for h in hareketler])
            
            st.dataframe(df_h, use_container_width=True, hide_index=True)
            
            # Excel raporu
            if st.button("📊 Excel'e Aktar"):
                output = io.BytesIO()
                wb = Workbook()
                ws = wb.active
                ws.title = "Hareketler"
                headers = list(df_h.columns)
                ws.append(headers)
                for row in df_h.values.tolist():
                    ws.append([str(c) for c in row])
                for i, col in enumerate(headers, 1):
                    ws.column_dimensions[chr(64+i)].width = 18
                wb.save(output)
                output.seek(0)
                st.download_button(
                    "⬇️ Excel İndir",
                    data=output.getvalue(),
                    file_name=f"elfiga_hareketler_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:
            st.info("Hareket kaydı bulunamadı.")
    
    with tab3:
        st.markdown("#### 🚚 Sevkiyat Yönetimi")
        
        # Üretim kayıtlarını listele
        uretim_kayitlari = [h for h in v["hareketler"] if h.get("islem") == "Üretim"]
        
        if not uretim_kayitlari:
            st.info("Sevkiyat eklemek için önce üretim yapılmalıdır.")
            return
        
        secenekler = [f"{h['parti']} - {h['malzeme']} - {h['miktar']} KG - {h['tarih']}" for h in uretim_kayitlari]
        secim = st.selectbox("Üretim Partisi Seç", secenekler)
        
        if secim:
            secilen_index = secenekler.index(secim)
            hedef_h = uretim_kayitlari[secilen_index]
            hedef_global_index = v["hareketler"].index(hedef_h)
            
            st.markdown(f"**Parti:** {hedef_h['parti']} | **Miktar:** {hedef_h['miktar']} KG")
            
            mevcut_sevk = hedef_h.get("sevkiyat_detay", [])
            
            # Sevkiyat ekleme formu
            with st.form("sevkiyat_form"):
                st.markdown("**Yeni Sevkiyat Ekle**")
                col_s1, col_s2, col_s3, col_s4 = st.columns(4)
                with col_s1:
                    s_tarih = st.date_input("Tarih", key="s_tarih")
                with col_s2:
                    s_firma = st.text_input("Firma Adı")
                with col_s3:
                    s_miktar = st.number_input("Miktar (KG)", min_value=0.0, step=0.5)
                with col_s4:
                    s_fatura = st.text_input("Fatura No")
                
                fire = st.slider("Fire Oranı (%)", 0, 20, int(float(hedef_h.get("fire_orani", 4))))
                notlar = st.text_area("Notlar", value=hedef_h.get("uretim_notu", ""))
                
                col_btn1, col_btn2 = st.columns(2)
                with col_btn1:
                    ekle = st.form_submit_button("➕ Sevkiyat Ekle", type="primary")
                with col_btn2:
                    kaydet = st.form_submit_button("💾 Kaydet")
                
                if ekle and s_firma and s_miktar > 0:
                    mevcut_sevk.append({
                        "tarih": s_tarih.strftime("%d.%m.%Y"),
                        "firma": s_firma, "miktar": s_miktar, "fatura": s_fatura
                    })
                    v["hareketler"][hedef_global_index]["sevkiyat_detay"] = mevcut_sevk
                    v["hareketler"][hedef_global_index]["fire_orani"] = str(fire)
                    v["hareketler"][hedef_global_index]["uretim_notu"] = notlar
                    veriler_kaydet(v)
                    st.success("✅ Sevkiyat eklendi!")
                    st.rerun()
                
                if kaydet:
                    v["hareketler"][hedef_global_index]["sevkiyat_detay"] = mevcut_sevk
                    v["hareketler"][hedef_global_index]["fire_orani"] = str(fire)
                    v["hareketler"][hedef_global_index]["uretim_notu"] = notlar
                    veriler_kaydet(v)
                    st.success("✅ Kaydedildi!")
            
            # Mevcut sevkiyatlar
            if mevcut_sevk:
                st.markdown("**Mevcut Sevkiyatlar:**")
                df_sevk = pd.DataFrame(mevcut_sevk)
                st.dataframe(df_sevk, use_container_width=True, hide_index=True)
                
                toplam_sevk = sum(float(s["miktar"]) for s in mevcut_sevk)
                net = hedef_h["miktar"] * (1 - float(hedef_h.get("fire_orani", 4))/100)
                st.metric("Toplam Sevkiyat", f"{toplam_sevk:.2f} KG")
                st.metric("Net Dağıtılabilir", f"{net:.2f} KG")
                
                # Rapor HTML
                if st.button("🖨️ Sevkiyat Raporu Oluştur"):
                    html = f"""
                    <html><head><style>
                    body{{font-family:Arial,sans-serif;font-size:13px;padding:20px}}
                    h2{{color:#007AFF;border-bottom:2px solid #007AFF;padding-bottom:8px}}
                    table{{width:100%;border-collapse:collapse;margin:15px 0}}
                    th,td{{border:1px solid #ddd;padding:8px;text-align:left}}
                    th{{background:#f5f5f7;font-weight:bold}}
                    .box{{background:#f9f9f9;border:1px solid #ddd;padding:12px;border-radius:8px;margin:10px 0}}
                    </style></head><body>
                    <h2>🥟 ELFİGA MANTI - Sevkiyat Raporu</h2>
                    <div class='box'>
                        <b>Ürün:</b> {hedef_h['malzeme']}<br>
                        <b>Parti No:</b> {hedef_h['parti']}<br>
                        <b>Üretim Miktarı:</b> {hedef_h['miktar']} KG<br>
                        <b>Fire Oranı:</b> %{hedef_h.get('fire_orani','4')}<br>
                        <b>Net Dağıtılabilir:</b> {net:.2f} KG<br>
                        <b>Not:</b> {notlar}
                    </div>
                    <h3>Sevkiyat Detayları</h3>
                    <table>
                        <tr><th>Tarih</th><th>Firma</th><th>Miktar (KG)</th><th>Fatura</th></tr>
                        {"".join(f"<tr><td>{s['tarih']}</td><td>{s['firma']}</td><td>{s['miktar']}</td><td>{s['fatura']}</td></tr>" for s in mevcut_sevk)}
                        <tr><td colspan='2'><b>TOPLAM</b></td><td><b>{toplam_sevk:.2f}</b></td><td></td></tr>
                    </table>
                    <br><br>
                    <div style='display:flex;justify-content:space-between;margin-top:60px'>
                        <div style='border-top:1px solid #000;padding-top:8px;width:200px;text-align:center'>Üretim Sorumlusu</div>
                        <div style='border-top:1px solid #000;padding-top:8px;width:200px;text-align:center'>Onay</div>
                    </div>
                    </body></html>
                    """
                    st.download_button(
                        "⬇️ Raporu İndir (HTML)",
                        data=html.encode("utf-8"),
                        file_name=f"sevkiyat_{hedef_h['parti']}.html",
                        mime="text/html"
                    )

# ─────────────────────────────────────────────
# SAYFA: AYARLAR
# ─────────────────────────────────────────────
def ayarlar_sayfasi(v):
    st.markdown("## ⚙️ Ayarlar")
    
    tab1, tab2, tab3 = st.tabs(["📋 Reçete Yönetimi", "👤 Kullanıcı Ayarları", "🔧 Sistem"])
    
    with tab1:
        col1, col2 = st.columns([1, 2])
        
        with col1:
            st.markdown("#### Reçeteler")
            recete_keys = list(v["coklu_receteler"].keys())
            secili_recete = st.selectbox("Reçete Seç", recete_keys)
            
            # Yeni reçete
            with st.expander("➕ Yeni Reçete Ekle"):
                yeni_ad = st.text_input("Reçete Adı")
                if st.button("Ekle (Aktif reçeteden kopyala)"):
                    if yeni_ad and yeni_ad not in v["coklu_receteler"]:
                        aktif_data = v["coklu_receteler"][v["aktif_recete_adi"]]
                        v["coklu_receteler"][yeni_ad] = {
                            "oranlar": aktif_data["oranlar"].copy(),
                            "etiket_tipi": aktif_data["etiket_tipi"]
                        }
                        v["aktif_recete_adi"] = yeni_ad
                        veriler_kaydet(v)
                        st.success(f"✅ '{yeni_ad}' eklendi!")
                        st.rerun()
            
            # Etiket tipi
            mevcut_tip = v["coklu_receteler"][secili_recete].get("etiket_tipi", "Soyalı")
            yeni_tip = st.selectbox("Etiket Tipi", ["Soyalı", "Soyasız"],
                                    index=0 if mevcut_tip == "Soyalı" else 1)
            if st.button("Etiket Tipini Kaydet"):
                v["coklu_receteler"][secili_recete]["etiket_tipi"] = yeni_tip
                veriler_kaydet(v)
                st.success("✅ Kaydedildi!")
            
            # Reçete sil
            if len(recete_keys) > 1:
                with st.expander("🗑️ Reçete Sil"):
                    silinecek = st.selectbox("Silinecek Reçete", [r for r in recete_keys if r != v["aktif_recete_adi"]])
                    if st.button("Sil", type="secondary"):
                        del v["coklu_receteler"][silinecek]
                        veriler_kaydet(v)
                        st.success(f"✅ '{silinecek}' silindi!")
                        st.rerun()
        
        with col2:
            st.markdown(f"#### {secili_recete} - Oranlar")
            recete_oranlari = v["coklu_receteler"][secili_recete]["oranlar"]
            toplam_oran = sum(recete_oranlari.values())
            
            if abs(toplam_oran - 1.0) > 0.01:
                st.warning(f"⚠️ Toplam oran: {toplam_oran:.4f} (1.0 olmalı!)")
            else:
                st.success(f"✅ Toplam oran: {toplam_oran:.4f}")
            
            df_rec = pd.DataFrame([
                {"Malzeme": m, "Oran": o, "Yüzde (%)": f"{o*100:.2f}%"}
                for m, o in recete_oranlari.items()
            ])
            st.dataframe(df_rec, use_container_width=True, hide_index=True)
            
            # Oran güncelle
            with st.expander("✏️ Oran Güncelle"):
                malzeme_sec = st.selectbox("Malzeme", list(recete_oranlari.keys()))
                yeni_oran = st.number_input("Yeni Oran (0-1 arası)",
                                           min_value=0.0, max_value=1.0, step=0.001,
                                           value=float(recete_oranlari.get(malzeme_sec, 0)))
                if st.button("Güncelle", type="primary"):
                    v["coklu_receteler"][secili_recete]["oranlar"][malzeme_sec] = yeni_oran
                    veriler_kaydet(v)
                    st.success(f"✅ {malzeme_sec} oranı güncellendi!")
                    st.rerun()
    
    with tab2:
        st.markdown("#### Parola Değiştir")
        with st.form("parola_form"):
            eski_parola = st.text_input("Mevcut Parola", type="password")
            yeni_parola = st.text_input("Yeni Parola", type="password")
            yeni_tekrar = st.text_input("Yeni Parola (Tekrar)", type="password")
            
            if st.form_submit_button("Parolayı Değiştir", type="primary"):
                user = kullanici_dogrula(st.session_state.username, eski_parola)
                if not user:
                    st.error("❌ Mevcut parola hatalı!")
                elif yeni_parola != yeni_tekrar:
                    st.error("❌ Yeni parolalar eşleşmiyor!")
                elif len(yeni_parola) < 6:
                    st.error("❌ Parola en az 6 karakter olmalı!")
                else:
                    conn = get_db()
                    cur = conn.cursor()
                    cur.execute("UPDATE kullanicilar SET password_hash = %s WHERE username = %s",
                               (hash_password(yeni_parola), st.session_state.username))
                    cur.close()
                    st.success("✅ Parola değiştirildi!")
        
        if st.session_state.get("role") == "admin":
            st.markdown("---")
            st.markdown("#### 👤 Yeni Kullanıcı Ekle (Admin)")
            with st.form("yeni_kullanici"):
                k_ad = st.text_input("Kullanıcı Adı")
                k_parola = st.text_input("Parola", type="password")
                k_rol = st.selectbox("Rol", ["user", "admin"])
                if st.form_submit_button("Kullanıcı Ekle"):
                    if k_ad and k_parola:
                        conn = get_db()
                        cur = conn.cursor()
                        try:
                            cur.execute("INSERT INTO kullanicilar (username, password_hash, role) VALUES (%s, %s, %s)",
                                       (k_ad, hash_password(k_parola), k_rol))
                            cur.close()
                            st.success(f"✅ '{k_ad}' kullanıcısı eklendi!")
                        except:
                            st.error("Bu kullanıcı adı zaten mevcut!")
    
    with tab3:
        st.markdown("#### Sistem Bilgileri")
        st.info(f"""
        - **Toplam Hareket:** {len(v['hareketler'])}
        - **Toplam Malzeme:** {len(v['stoklar'])}
        - **Toplam Reçete:** {len(v['coklu_receteler'])}
        - **Detaylı Stok Kaydı:** {len(v['detayli_stok'])}
        - **Aktif Reçete:** {v['aktif_recete_adi']}
        """)
        
        st.markdown("---")
        
        # JSON yedek indir
        if st.button("📥 Tüm Veriyi JSON Olarak İndir"):
            st.download_button(
                "⬇️ Yedek İndir",
                data=json.dumps(v, ensure_ascii=False, indent=2).encode("utf-8"),
                file_name=f"elfiga_yedek_{datetime.now().strftime('%Y%m%d_%H%M')}.json",
                mime="application/json"
            )
        
        # JSON yükle
        st.markdown("**📤 JSON Yedekten Geri Yükle**")
        yuklenen = st.file_uploader("JSON Dosyası Seç", type=["json"])
        if yuklenen:
            if st.button("⚠️ Geri Yükle (Mevcut veriler silinir!)", type="secondary"):
                try:
                    icerik = yuklenen.read()
                    yeni_veri = json.loads(icerik.decode('utf-8'))
                    
                    # Eksik alanları doldur
                    for h in yeni_veri.get('hareketler', []):
                        if 'fiyat' not in h: h['fiyat'] = 0
                        if 'fatura' not in h: h['fatura'] = '-'
                        if 'parti' not in h: h['parti'] = '-'
                        if 'islem' not in h: h['islem'] = 'Giriş'
                    
                    for s in yeni_veri.get('detayli_stok', []):
                        if 'kalan' not in s: s['kalan'] = s.get('miktar', 0)
                        if 'fatura' not in s: s['fatura'] = '-'
                        if 'parti' not in s: s['parti'] = '-'
                    
                    # DB'ye kaydet
                    basarili = veriler_kaydet(yeni_veri)
                    if basarili:
                        # Session'ı temizle, DB'den taze yüklensin
                        st.session_state.veriler = yeni_veri
                        st.success(f"✅ {len(yeni_veri.get('hareketler',[]))} hareket, {len(yeni_veri.get('detayli_stok',[]))} stok kaydı yüklendi!")
                        st.rerun()
                    else:
                        st.error("❌ Kaydetme başarısız!")
                except Exception as e:
                    st.error(f"❌ Hata: {e}")

# ─────────────────────────────────────────────
# SICAKLIK YARDIMCI FONKSİYONLAR
# ─────────────────────────────────────────────
def sicaklik_db_init():
    """Sıcaklık tablolarını oluştur (app.py'nin init_db() fonksiyonuna ekle)"""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS sicaklik_sensorler (
            id SERIAL PRIMARY KEY,
            tenant_id INTEGER,
            sensor_id TEXT NOT NULL,
            sensor_adi TEXT DEFAULT '',
            konum TEXT DEFAULT '',
            min_alarm NUMERIC DEFAULT -25.0,
            max_alarm NUMERIC DEFAULT 5.0,
            aktif BOOLEAN DEFAULT TRUE,
            api_key TEXT,
            olusturma_tarihi TIMESTAMP DEFAULT NOW(),
            UNIQUE(tenant_id, sensor_id)
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS sicaklik_olcumler (
            id SERIAL PRIMARY KEY,
            tenant_id INTEGER,
            sensor_id TEXT NOT NULL,
            sicaklik NUMERIC NOT NULL,
            nem NUMERIC,
            kayit_zamani TIMESTAMP DEFAULT NOW()
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS sicaklik_alarmlar (
            id SERIAL PRIMARY KEY,
            tenant_id INTEGER,
            sensor_id TEXT,
            sicaklik NUMERIC,
            alarm_tipi TEXT,
            mesaj TEXT,
            goruldu BOOLEAN DEFAULT FALSE,
            kayit_zamani TIMESTAMP DEFAULT NOW()
        )
    """)
    cur.close()

def sensorleri_getir(tenant_id):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM sicaklik_sensorler WHERE tenant_id = %s ORDER BY sensor_id", (tenant_id,))
    rows = cur.fetchall()
    cur.close()
    return [dict(r) for r in rows]

def sensor_ekle(tenant_id, sensor_id, sensor_adi, konum, min_al, max_al):
    import secrets
    api_key = "SK-" + secrets.token_hex(16)
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO sicaklik_sensorler (tenant_id, sensor_id, sensor_adi, konum, min_alarm, max_alarm, api_key)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT (tenant_id, sensor_id) DO UPDATE 
        SET sensor_adi=%s, konum=%s, min_alarm=%s, max_alarm=%s
    """, (tenant_id, sensor_id, sensor_adi, konum, min_al, max_al, api_key,
          sensor_adi, konum, min_al, max_al))
    cur.close()
    return api_key

def son_olcumleri_getir(tenant_id, sensor_idler):
    """Her sensörün son ölçümünü getir"""
    if not sensor_idler:
        return {}
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    sonuclar = {}
    for sid in sensor_idler:
        cur.execute("""
            SELECT sicaklik, nem, kayit_zamani 
            FROM sicaklik_olcumler 
            WHERE tenant_id = %s AND sensor_id = %s 
            ORDER BY kayit_zamani DESC LIMIT 1
        """, (tenant_id, sid))
        row = cur.fetchone()
        if row:
            sonuclar[sid] = dict(row)
    cur.close()
    return sonuclar

def gecmis_olcumleri_getir(tenant_id, sensor_id, limit=200):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT sicaklik, nem, kayit_zamani 
        FROM sicaklik_olcumler 
        WHERE tenant_id = %s AND sensor_id = %s 
        ORDER BY kayit_zamani DESC LIMIT %s
    """, (tenant_id, sensor_id, limit))
    rows = cur.fetchall()
    cur.close()
    return [dict(r) for r in reversed(rows)]  # Eski→yeni sırala

def aktif_alarmlari_getir(tenant_id):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT * FROM sicaklik_alarmlar 
        WHERE tenant_id = %s AND goruldu = FALSE 
        ORDER BY kayit_zamani DESC LIMIT 20
    """, (tenant_id,))
    rows = cur.fetchall()
    cur.close()
    return [dict(r) for r in rows]

def alarmi_goruldu_isaretle(tenant_id):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE sicaklik_alarmlar SET goruldu = TRUE WHERE tenant_id = %s", (tenant_id,))
    cur.close()



# ─────────────────────────────────────────────
# SICAKLIK TAKİP SAYFASI
# ─────────────────────────────────────────────
def sicaklik_sayfasi(v):
    st.markdown("## 🌡️ Sıcaklık Takip Sistemi")

    tenant_id = st.session_state.get("tenant_id", 0)

    # DB tablolarını hazırla
    try:
        sicaklik_db_init()
    except:
        pass

    # Aktif alarmlar
    try:
        alarmlar = aktif_alarmlari_getir(tenant_id)
        if alarmlar:
            for alarm in alarmlar[:3]:
                st.error(f"🚨 **ALARM** | {alarm['sensor_id']} | {alarm['mesaj']} | {str(alarm['kayit_zamani'])[:16]}")
            col_alarm = st.columns([3, 1])
            with col_alarm[1]:
                if st.button("✅ Tümünü Okundu İşaretle"):
                    alarmi_goruldu_isaretle(tenant_id)
                    st.rerun()
    except:
        pass

    tab1, tab2, tab3 = st.tabs(["📊 Canlı İzleme", "📈 Geçmiş & Grafik", "⚙️ Sensör Ayarları"])

    # ── CANLI İZLEME ────────────────────────────────────────────────────
    with tab1:
        st.markdown("#### 🔴 Canlı Sensör Durumu")
        st.caption("Sayfa her 30 saniyede otomatik güncellenir")

        try:
            sensorler = sensorleri_getir(tenant_id)
        except:
            sensorler = []

        if not sensorler:
            st.info("Henüz sensör tanımlanmamış. '⚙️ Sensör Ayarları' sekmesinden ekleyin.")
        else:
            son_olcumler = son_olcumleri_getir(tenant_id, [s["sensor_id"] for s in sensorler])

            # Sensör kartları
            cols = st.columns(min(len(sensorler), 3))
            for i, sensor in enumerate(sensorler):
                with cols[i % 3]:
                    olcum = son_olcumler.get(sensor["sensor_id"])
                    if olcum:
                        sicaklik = float(olcum["sicaklik"])
                        nem = float(olcum["nem"]) if olcum.get("nem") else None
                        zaman = str(olcum["kayit_zamani"])[:16]

                        # Renk belirle
                        min_al = float(sensor["min_alarm"])
                        max_al = float(sensor["max_alarm"])

                        if sicaklik < min_al:
                            renk = "#FF3B30"
                            durum = "⬇️ DÜŞÜK ALARM"
                        elif sicaklik > max_al:
                            renk = "#FF9500"
                            durum = "⬆️ YÜKSEK ALARM"
                        else:
                            renk = "#34C759"
                            durum = "✅ Normal"

                        st.markdown(f"""
<div style="background:#1a1a2e;border:2px solid {renk};border-radius:14px;padding:20px;margin-bottom:12px;text-align:center">
  <div style="color:#8E8E93;font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:1px">{sensor['sensor_adi'] or sensor['sensor_id']}</div>
  <div style="color:#8E8E93;font-size:10px;margin:2px 0 8px 0">📍 {sensor['konum'] or '-'}</div>
  <div style="color:{renk};font-size:48px;font-weight:800;line-height:1">{sicaklik:.1f}°C</div>
  {'<div style="color:#007AFF;font-size:16px;margin-top:6px">💧 ' + f'{nem:.1f}%' + '</div>' if nem else ''}
  <div style="color:{renk};font-size:12px;font-weight:600;margin-top:8px">{durum}</div>
  <div style="color:#555;font-size:10px;margin-top:4px">🕐 {zaman}</div>
  <div style="color:#444;font-size:10px;margin-top:2px">Sınır: {min_al}°C / {max_al}°C</div>
</div>
                        """, unsafe_allow_html=True)
                    else:
                        st.markdown(f"""
<div style="background:#1a1a2e;border:2px solid #3a3a4e;border-radius:14px;padding:20px;margin-bottom:12px;text-align:center">
  <div style="color:#8E8E93;font-size:11px;font-weight:600">{sensor['sensor_adi'] or sensor['sensor_id']}</div>
  <div style="color:#555;font-size:32px;margin:12px 0">📡</div>
  <div style="color:#555;font-size:13px">Veri bekleniyor...</div>
  <div style="color:#444;font-size:10px;margin-top:4px">Sensör henüz bağlanmadı</div>
</div>
                        """, unsafe_allow_html=True)

        # Otomatik yenileme
        st.markdown("---")
        col_r1, col_r2 = st.columns([3, 1])
        with col_r2:
            if st.button("🔄 Şimdi Yenile", use_container_width=True):
                st.rerun()
        with col_r1:
            st.caption(f"Son güncelleme: {datetime.now().strftime('%H:%M:%S')}")

    # ── GEÇMİŞ & GRAFİK ─────────────────────────────────────────────────
    with tab2:
        st.markdown("#### 📈 Sıcaklık Geçmişi")

        try:
            sensorler = sensorleri_getir(tenant_id)
        except:
            sensorler = []

        if not sensorler:
            st.info("Sensör tanımlanmamış.")
        else:
            sensor_secenekler = {f"{s['sensor_adi'] or s['sensor_id']} ({s['konum'] or '-'})": s["sensor_id"] for s in sensorler}
            secim = st.selectbox("Sensör Seç", list(sensor_secenekler.keys()))
            secilen_sensor_id = sensor_secenekler[secim]

            col_limit, col_grafik_tip = st.columns(2)
            with col_limit:
                limit = st.selectbox("Kaç ölçüm?", [50, 100, 200, 500], index=1)
            with col_grafik_tip:
                grafik_tip = st.selectbox("Grafik Tipi", ["Çizgi", "Alan"])

            olcumler = gecmis_olcumleri_getir(tenant_id, secilen_sensor_id, limit)

            if olcumler:
                df = pd.DataFrame(olcumler)
                df["kayit_zamani"] = pd.to_datetime(df["kayit_zamani"])
                df["sicaklik"] = df["sicaklik"].astype(float)

                # İstatistikler
                c1, c2, c3, c4 = st.columns(4)
                with c1:
                    st.metric("🌡️ Son Ölçüm", f"{df['sicaklik'].iloc[-1]:.1f}°C")
                with c2:
                    st.metric("📉 Min", f"{df['sicaklik'].min():.1f}°C")
                with c3:
                    st.metric("📈 Max", f"{df['sicaklik'].max():.1f}°C")
                with c4:
                    st.metric("📊 Ortalama", f"{df['sicaklik'].mean():.1f}°C")

                # Grafik
                sensor_bilgi = next((s for s in sensorler if s["sensor_id"] == secilen_sensor_id), {})
                min_al = float(sensor_bilgi.get("min_alarm", -25))
                max_al = float(sensor_bilgi.get("max_alarm", 5))

                if grafik_tip == "Alan":
                    fig = px.area(df, x="kayit_zamani", y="sicaklik",
                                 title=f"Sıcaklık Geçmişi — {secim}",
                                 color_discrete_sequence=["#007AFF"])
                else:
                    fig = px.line(df, x="kayit_zamani", y="sicaklik",
                                 title=f"Sıcaklık Geçmişi — {secim}",
                                 color_discrete_sequence=["#007AFF"])

                # Alarm çizgileri
                fig.add_hline(y=min_al, line_dash="dash", line_color="#FF3B30",
                             annotation_text=f"Min Alarm ({min_al}°C)")
                fig.add_hline(y=max_al, line_dash="dash", line_color="#FF9500",
                             annotation_text=f"Max Alarm ({max_al}°C)")

                fig.update_layout(
                    paper_bgcolor="#0F0F13",
                    plot_bgcolor="#1a1a2e",
                    font_color="white",
                    xaxis_title="Zaman",
                    yaxis_title="Sıcaklık (°C)",
                    height=400
                )
                fig.update_xaxes(gridcolor="#2a2a3e")
                fig.update_yaxes(gridcolor="#2a2a3e")
                st.plotly_chart(fig, use_container_width=True)

                # Nem grafiği
                if "nem" in df.columns and df["nem"].notna().any():
                    fig2 = px.line(df, x="kayit_zamani", y="nem",
                                  title="Nem Geçmişi (%)",
                                  color_discrete_sequence=["#34C759"])
                    fig2.update_layout(
                        paper_bgcolor="#0F0F13", plot_bgcolor="#1a1a2e",
                        font_color="white", height=250
                    )
                    st.plotly_chart(fig2, use_container_width=True)

                # Excel indir
                if st.button("📊 Excel'e Aktar"):
                    output = io.BytesIO()
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Sıcaklık Verileri"
                    ws.append(["Zaman", "Sıcaklık (°C)", "Nem (%)"])
                    for _, row in df.iterrows():
                        ws.append([str(row["kayit_zamani"]), float(row["sicaklik"]),
                                  float(row["nem"]) if pd.notna(row.get("nem")) else ""])
                    wb.save(output)
                    output.seek(0)
                    st.download_button("⬇️ İndir",
                                      data=output.getvalue(),
                                      file_name=f"sicaklik_{secilen_sensor_id}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                                      mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            else:
                st.info("Bu sensörden henüz ölçüm gelmedi.")

    # ── SENSÖR AYARLARI ──────────────────────────────────────────────────
    with tab3:
        st.markdown("#### ⚙️ Sensör Tanımla / Düzenle")

        with st.form("sensor_ekle_form"):
            col1, col2 = st.columns(2)
            with col1:
                sensor_id = st.text_input("Sensör ID *", placeholder="don_odasi_1",
                                         help="Küçük harf, alt çizgi. Örn: don_odasi_1")
                sensor_adi = st.text_input("Sensör Adı", placeholder="Dondurma Odası 1")
                konum = st.text_input("Konum", placeholder="B Blok, 2. Kat")
            with col2:
                min_alarm = st.number_input("Min Alarm (°C)", value=-25.0, step=0.5,
                                           help="Bu sıcaklığın altına inince alarm ver")
                max_alarm = st.number_input("Max Alarm (°C)", value=5.0, step=0.5,
                                           help="Bu sıcaklığın üstüne çıkınca alarm ver")

            kaydet = st.form_submit_button("💾 Sensörü Kaydet", type="primary")

            if kaydet:
                if not sensor_id:
                    st.error("Sensör ID zorunlu!")
                else:
                    try:
                        api_key = sensor_ekle(tenant_id, sensor_id, sensor_adi, konum, min_alarm, max_alarm)
                        st.success(f"✅ Sensör kaydedildi!")
                        st.code(f"""
// ESP32 Arduino Kodu (sicaklik_api.py'ye gönderir)
// Bu bilgileri Arduino koduna ekle:

const char* API_URL = "{SICAKLIK_API_URL}/api/sicaklik";
const char* API_KEY = "{api_key}";

// HTTP POST örneği:
// POST {SICAKLIK_API_URL}/api/sicaklik
// Header: X-API-Key: {api_key}
// Body: {{"sensor_id": "{sensor_id}", "sicaklik": 23.5, "nem": 65.0}}
                        """, language="cpp")
                    except Exception as e:
                        st.error(f"Hata: {e}")

        st.markdown("---")
        st.markdown("#### 📋 Kayıtlı Sensörler")

        try:
            sensorler = sensorleri_getir(tenant_id)
            if sensorler:
                df_s = pd.DataFrame([{
                    "Sensör ID": s["sensor_id"],
                    "Adı": s["sensor_adi"],
                    "Konum": s["konum"],
                    "Min Alarm (°C)": s["min_alarm"],
                    "Max Alarm (°C)": s["max_alarm"],
                    "Aktif": "✅" if s["aktif"] else "❌",
                    "API Key": s["api_key"]
                } for s in sensorler])
                st.dataframe(df_s, use_container_width=True, hide_index=True)
            else:
                st.info("Henüz sensör eklenmemiş.")
        except Exception as e:
            st.info("Sensör tablosu hazırlanıyor...")

        # ESP32 örnek kod
        st.markdown("---")
        st.markdown("#### 📟 ESP32 Örnek Arduino Kodu")
        st.code("""
#include <WiFi.h>
#include <HTTPClient.h>
#include <DHT.h>

// WiFi Ayarları
const char* WIFI_SSID     = "WIFI_ADINIZ";
const char* WIFI_PASSWORD = "WIFI_SIFRENIZ";

// API Ayarları (Sensör ekledikten sonra buraya kopyala)
const char* API_URL = "https://SIZIN-API-URL.railway.app/api/sicaklik";
const char* API_KEY = "SK-xxxxxxxxxxxxxxxxxxxx";

// DHT Sensör
#define DHT_PIN 4
#define DHT_TYPE DHT22
DHT dht(DHT_PIN, DHT_TYPE);

void setup() {
  Serial.begin(115200);
  dht.begin();
  
  WiFi.begin(WIFI_SSID, WIFI_PASSWORD);
  Serial.print("WiFi bağlanıyor");
  while (WiFi.status() != WL_CONNECTED) {
    delay(500);
    Serial.print(".");
  }
  Serial.println(" Bağlandı!");
}

void loop() {
  float sicaklik = dht.readTemperature();
  float nem      = dht.readHumidity();
  
  if (isnan(sicaklik) || isnan(nem)) {
    Serial.println("Sensör okuma hatası!");
    delay(5000);
    return;
  }
  
  Serial.printf("Sıcaklık: %.1f°C  Nem: %.1f%%\\n", sicaklik, nem);
  
  if (WiFi.status() == WL_CONNECTED) {
    HTTPClient http;
    http.begin(API_URL);
    http.addHeader("Content-Type", "application/json");
    http.addHeader("X-API-Key", API_KEY);
    
    String body = "{\\"sicaklik\\":" + String(sicaklik, 1) + 
                  ",\\"nem\\":" + String(nem, 1) + "}";
    
    int httpCode = http.POST(body);
    
    if (httpCode == 200) {
      Serial.println("Veri gönderildi ✓");
    } else {
      Serial.printf("HTTP Hata: %d\\n", httpCode);
    }
    http.end();
  }
  
  delay(30000); // 30 saniyede bir gönder
}
        """, language="cpp")


# ─────────────────────────────────────────────
# ANA UYGULAMA
# ─────────────────────────────────────────────
def main():
    # DB başlat
    init_db()
    admin_kullanici_olustur()
    
    # Session state başlat
    if "logged_in" not in st.session_state:
        st.session_state.logged_in = False
    
    # Login kontrol
    if not st.session_state.logged_in:
        login_sayfasi()
        return
    
    # Verileri her zaman DB'den taze yükle
    v = veriler_yukle()
    st.session_state.veriler = v
    
    # Menü
    menu = sidebar_menu()
    
    # Sayfa yönlendirme
    if menu == "📊 Patron Ekranı":
        patron_ekrani(v)
    elif menu == "🏭 Üretim":
        uretim_sayfasi(v)
    elif menu == "📍 Depom":
        depom_sayfasi(v)
    elif menu == "📦 Depo & Giriş":
        depo_giris_sayfasi(v)
    elif menu == "🌡️ Sıcaklık":
        sicaklik_sayfasi(v)
    elif menu == "⚙️ Ayarlar":
        ayarlar_sayfasi(v)

if __name__ == "__main__":
    main()

